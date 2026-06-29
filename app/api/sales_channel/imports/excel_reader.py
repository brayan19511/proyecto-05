"""Read small XLSX files into normalized SKU rows without storing the file."""

from dataclasses import dataclass
from hashlib import sha256
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.api.sales_channel.imports.schemas import (
    SkuImportIssue,
    SkuImportMode,
)
from app.core.exceptions import ValidationError


MAX_XLSX_BYTES = 50 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000

SKU_HEADERS = {"sku"}
ACTIVE_HEADERS = {"active", "is_active", "on", "on_off", "on/off"}


@dataclass(frozen=True)
class ParsedSkuWorkbook:
    filename: str
    sha256: str
    received: int
    rows: list[dict]
    errors: list[SkuImportIssue]


def _normalize_header(value) -> str:
    return str(value or "").strip().casefold().replace(" ", "_")


def _normalize_sku(value) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    normalized = str(value).strip()
    return normalized or None


def _parse_active(value) -> bool | None:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in (0, 1):
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().casefold()
        if normalized in {"1", "true", "on", "yes", "si", "activo"}:
            return True
        if normalized in {"0", "false", "off", "no", "inactivo"}:
            return False
    return None


def _read_file_bytes(file: BinaryIO) -> bytes:
    file.seek(0)
    content = file.read(MAX_XLSX_BYTES + 1)
    file.seek(0)
    if len(content) > MAX_XLSX_BYTES:
        raise ValidationError("El archivo supera el limite de 5 MB")
    return content


def parse_sku_workbook(
    upload: UploadFile,
    mode: SkuImportMode,
) -> ParsedSkuWorkbook:
    """Validate headers and return rows ready for the domain service."""
    filename = Path(upload.filename or "import.xlsx").name
    if Path(filename).suffix.casefold() != ".xlsx":
        raise ValidationError("Solo se permiten archivos con extension .xlsx")

    content = _read_file_bytes(upload.file)
    try:
        workbook = load_workbook(
            upload.file,
            read_only=True,
            data_only=True,
        )
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ValidationError("El archivo XLSX no es valido") from exc

    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        if header_values is None:
            raise ValidationError("El archivo no contiene encabezados")

        headers = [_normalize_header(value) for value in header_values]
        sku_index = next(
            (index for index, value in enumerate(headers) if value in SKU_HEADERS),
            None,
        )
        active_index = next(
            (
                index
                for index, value in enumerate(headers)
                if value in ACTIVE_HEADERS
            ),
            None,
        )
        errors: list[SkuImportIssue] = []
        if sku_index is None:
            errors.append(
                SkuImportIssue(
                    row=1,
                    field="sku",
                    message="Falta la columna sku",
                )
            )
        if mode == SkuImportMode.STATUS_UPDATE and active_index is None:
            errors.append(
                SkuImportIssue(
                    row=1,
                    field="active",
                    message="Falta la columna active u on/off",
                )
            )
        if errors:
            return ParsedSkuWorkbook(
                filename=filename,
                sha256=sha256(content).hexdigest(),
                received=0,
                rows=[],
                errors=errors,
            )

        rows: list[dict] = []
        seen_skus: dict[str, int] = {}
        received = 0
        for row_number, values in enumerate(iterator, start=2):
            if not any(value not in (None, "") for value in values):
                continue
            received += 1
            if received > MAX_IMPORT_ROWS:
                errors.append(
                    SkuImportIssue(
                        row=row_number,
                        message="El archivo supera el limite de 20000 filas",
                    )
                )
                break

            sku_value = values[sku_index] if sku_index < len(values) else None
            sku = _normalize_sku(sku_value)
            if not sku or len(sku) > 255:
                errors.append(
                    SkuImportIssue(
                        row=row_number,
                        field="sku",
                        message="El SKU debe contener entre 1 y 255 caracteres",
                    )
                )
                continue

            normalized_sku = sku.casefold()
            if normalized_sku in seen_skus:
                errors.append(
                    SkuImportIssue(
                        row=row_number,
                        field="sku",
                        message=(
                            f"SKU duplicado; primera aparicion en fila "
                            f"{seen_skus[normalized_sku]}"
                        ),
                    )
                )
                continue
            seen_skus[normalized_sku] = row_number

            row = {"sku": sku}
            if mode == SkuImportMode.STATUS_UPDATE:
                active_value = (
                    values[active_index]
                    if active_index < len(values)
                    else None
                )
                active = _parse_active(active_value)
                if active is None:
                    errors.append(
                        SkuImportIssue(
                            row=row_number,
                            field="active",
                            message="Estado invalido; use on/off, true/false o 1/0",
                        )
                    )
                    continue
                row["active"] = active
            rows.append(row)

        if received == 0:
            errors.append(
                SkuImportIssue(message="El archivo no contiene filas de datos")
            )

        return ParsedSkuWorkbook(
            filename=filename,
            sha256=sha256(content).hexdigest(),
            received=received,
            rows=rows,
            errors=errors,
        )
    finally:
        workbook.close()
        upload.file.seek(0)
