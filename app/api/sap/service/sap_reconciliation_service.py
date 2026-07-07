from collections import defaultdict
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.exceptions import ValidationError


REQUIRED_COLUMNS = {
    "ReconNum",
    "SrcObjAbs",
    "SrcObjTyp",
    "TransRowId",
    "TransId",
    "ShortName",
    "ReconcileAmount",
}
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024


class SapReconciliationExcelService:
    """Convierte un Excel de conciliacion en payloads listos para SAP."""

    def build_payloads(
        self,
        file: BinaryIO,
        *,
        filename: str | None,
        recon_date: date,
    ) -> dict[str, dict]:
        content = self._read_file(file, filename)
        worksheet = self._open_first_sheet(content)
        headers = self._read_headers(worksheet)
        groups = self._group_rows_by_recon_num(worksheet, headers)
        return self._build_payloads(groups, recon_date)

    def _read_file(self, file: BinaryIO, filename: str | None) -> bytes:
        # Validamos extension y tamano antes de abrir el archivo para fallar
        # rapido si el usuario sube algo distinto a la plantilla esperada.
        if not filename or not filename.lower().endswith(".xlsx"):
            raise ValidationError("Solo se permiten archivos .xlsx")

        content = file.read(MAX_FILE_SIZE_BYTES + 1)
        if len(content) > MAX_FILE_SIZE_BYTES:
            raise ValidationError("El archivo supera el limite de 50 MB")
        if not content:
            raise ValidationError("El archivo esta vacio")
        return content

    def _open_first_sheet(self, content: bytes):
        try:
            # data_only=True lee el valor calculado de celdas con formulas.
            workbook = load_workbook(
                BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except (InvalidFileException, OSError, ValueError) as exc:
            raise ValidationError("El archivo XLSX no es valido") from exc
        return workbook.active

    def _read_headers(self, worksheet) -> dict[str, int]:
        row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row:
            raise ValidationError("El archivo no contiene encabezados")

        headers = {
            str(value).strip(): index
            for index, value in enumerate(row)
            if value is not None and str(value).strip()
        }
        missing = sorted(REQUIRED_COLUMNS - set(headers))
        if missing:
            raise ValidationError(
                "Faltan columnas obligatorias: " + ", ".join(missing)
            )
        return headers

    def _group_rows_by_recon_num(
        self,
        worksheet,
        headers: dict[str, int],
    ) -> dict[str, list[dict]]:
        groups: dict[str, list[dict]] = defaultdict(list)
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            recon_num = self._required(row, headers, "ReconNum", row_number)
            reconcile_amount = self._required(
                row,
                headers,
                "ReconcileAmount",
                row_number,
                parser=Decimal,
                type_name="numerico",
            )
            if reconcile_amount == 0:
                raise ValidationError(f"Fila {row_number}: ReconcileAmount no puede ser 0")

            # SAP recibe montos absolutos y el lado contable se define con
            # codDebit/codCredit segun el signo del ReconcileAmount en el Excel.
            groups[recon_num].append(
                {
                    "CashDiscount": None,
                    "CreditOrDebit": "codDebit" if reconcile_amount > 0 else "codCredit",
                    "ReconcileAmount": float(abs(reconcile_amount)),
                    "Selected": "tYES",
                    "ShortName": self._required(
                        row,
                        headers,
                        "ShortName",
                        row_number,
                    ),
                    "SrcObjAbs": self._required(
                        row,
                        headers,
                        "SrcObjAbs",
                        row_number,
                        parser=int,
                        type_name="numerico",
                    ),
                    "SrcObjTyp": self._required(
                        row,
                        headers,
                        "SrcObjTyp",
                        row_number,
                    ),
                    "TransId": self._required(
                        row,
                        headers,
                        "TransId",
                        row_number,
                        parser=int,
                        type_name="numerico",
                    ),
                    "TransRowId": self._required(
                        row,
                        headers,
                        "TransRowId",
                        row_number,
                        parser=int,
                        type_name="numerico",
                    ),
                    "_reconcile_amount": reconcile_amount,
                }
            )

        if not groups:
            raise ValidationError("El archivo no contiene filas para conciliar")
        return groups

    def _build_payloads(
        self,
        groups: dict[str, list[dict]],
        recon_date: date,
    ) -> dict[str, dict]:
        payloads = {}
        for recon_num, rows in groups.items():
            total = sum(row["_reconcile_amount"] for row in rows)
            if abs(total) > Decimal("0.01"):
                raise ValidationError(
                    f"ReconNum {recon_num}: la suma ReconcileAmount debe ser 0"
                )

            # SAP no necesita el ReconcileAmount original; solo el lado debito/credito
            # y el monto absoluto que se va a reconciliar.
            clean_rows = []
            for row in rows:
                clean = dict(row)
                clean.pop("_reconcile_amount", None)
                clean_rows.append(clean)

            payloads[recon_num] = {
                "CardOrAccount": "coaCard",
                "InternalReconciliationOpenTransRows": clean_rows,
                "ReconDate": recon_date.isoformat(),
            }
        return payloads

    def _required(
        self,
        row,
        headers: dict[str, int],
        column: str,
        row_number: int,
        *,
        parser=str,
        type_name: str = "texto",
    ):
        value = self._value(row, headers, column)
        if value is None or not str(value).strip():
            raise ValidationError(f"Fila {row_number}: {column} es obligatorio")
        try:
            parsed = parser(str(value).strip()) if parser is Decimal else parser(value)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ValidationError(
                f"Fila {row_number}: {column} debe ser {type_name}"
            ) from exc
        return parsed.strip() if isinstance(parsed, str) else parsed

    @staticmethod
    def _value(row, headers: dict[str, int], column: str):
        index = headers[column]
        return row[index] if index < len(row) else None
