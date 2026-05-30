# app/api/sap/finance/sap_finance_service.py
import json
import os
import io

import openpyxl
import pandas as pd
import numpy as np
from app.api.sap.finance.sap_finance_repository import SapFinanceRepository


class SapFinanceService:
    def __init__(self, db):
        self.db = db
        self.company = "SBO_RASH_PRODUCCION"
        self.repository = SapFinanceRepository(db, self.company)

    def _obtener_reglas_mock(self):
        """Simula tu tabla #TMP_Reglas_Gastos"""
        ruta_archivo = "app/api/sap/tmp/reglas.json"

        if not os.path.exists(ruta_archivo):
            print(
                f"Advertencia: El archivo {ruta_archivo} no existe. Devolviendo lista vacía."
            )
            return []

        with open(ruta_archivo, "r", encoding="utf-8") as archivo:
            data_completa = json.load(archivo)

        if isinstance(data_completa, dict):
            return data_completa.get("reglas_gastos", [])

        return data_completa

    def evaluar_regla_para_fila(self, fila, df_reglas):
        """Simula el TOP 1 y los WHERE del OUTER APPLY de SQL"""
        if df_reglas.empty:
            return pd.Series([None, "SIN CLASIFICAR", "SIN CODIGO", "SIN SUBCODIGO"])

        # Iteramos directamente sobre df_reglas (que ya viene ordenado por prioridad desde afuera)
        for _, regla in df_reglas.iterrows():
            # 1. Validar Tipo de Regla y Cuenta
            if (
                regla["tipo_regla"] == "CUENTA"
                and regla["cuenta"] != fila["cuenta_asociada"]
            ):
                continue
            if (
                regla["tipo_regla"] == "MIXTA"
                and regla["cuenta"] != fila["cuenta_asociada"]
            ):
                continue

            # 2. Filtro de Contrapartida y Centro de Costo
            if (
                pd.notna(regla["cuenta_contrapartida"])
                and regla["cuenta_contrapartida"] != fila["cuenta_contrapartida"]
            ):
                continue
            if (
                pd.notna(regla["centro_costo"])
                and regla["centro_costo"] != fila["centro_costo"]
            ):
                continue

            # 3. Filtros de Texto (Incluido)
            if pd.notna(regla["filtro_texto"]):
                texto_buscar = str(regla["filtro_texto"]).lower()
                prov = str(fila.get("proveedor") or "").lower()
                ref1 = str(fila.get("referencia_1") or "").lower()
                ref2 = str(fila.get("referencia_2") or "").lower()
                desc = str(fila.get("descripcion") or "").lower()

                if not (
                    texto_buscar in prov
                    or texto_buscar in ref1
                    or texto_buscar in ref2
                    or texto_buscar in desc
                ):
                    continue

            # 4. Filtros de Texto (Excluido)
            if pd.notna(regla["texto_excluido"]):
                texto_excluir = str(regla["texto_excluido"]).lower()
                prov = str(fila.get("proveedor") or "").lower()
                ref1 = str(fila.get("referencia_1") or "").lower()

                if texto_excluir in prov or texto_excluir in ref1:
                    continue

            # 5. Filtros de Monto
            monto = float(fila["cargo_abono_ml"] or 0)
            if pd.notna(regla["monto_min"]) and monto < float(regla["monto_min"]):
                continue
            if pd.notna(regla["monto_max"]) and monto > float(regla["monto_max"]):
                continue

            # SI CUMPLIÓ: Retornamos las 4 columnas correspondientes a la regla
            return pd.Series(
                [
                    regla["id_regla"],
                    regla["nombre_cuenta"],
                    regla["codigo"],
                    regla["subcodigo"],
                ]
            )

        # Si no cumplió ninguna regla del bucle
        return pd.Series([None, "SIN CLASIFICAR", "SIN CODIGO", "SIN SUBCODIGO"])

    def get_libro_mayor_account(self, start_date, end_date, account):
        # 1. Obtener la data cruda de SAP
        raw_data = self.repository.get_libro_mayor_by_account(
            start_date, end_date, account
        )
        if not raw_data:
            return []

        df_sap = pd.DataFrame([dict(row) for row in raw_data])

        # Cargar y ordenar las reglas una sola vez aquí afuera
        reglas_list = self._obtener_reglas_mock()
        df_reglas = pd.DataFrame(reglas_list)

        # =========================================================================
        # PROTECCIÓN CLAVE: Aseguramos que existan todas las columnas en df_reglas
        # para evitar el error 'KeyError' si el JSON viene incompleto
        # =========================================================================
        columnas_obligatorias = [
            "id_regla",
            "tipo_regla",
            "cuenta",
            "centro_costo",
            "cuenta_contrapartida",
            "filtro_texto",
            "texto_excluido",
            "monto_min",
            "monto_max",
            "nombre_cuenta",
            "codigo",
            "subcodigo",
            "prioridad",
        ]

        for col in columnas_obligatorias:
            if col not in df_reglas.columns:
                df_reglas[col] = None  # Si no existe en el JSON, la creamos vacía

        if not df_reglas.empty and "prioridad" in df_reglas.columns:
            df_reglas = df_reglas.sort_values(by="prioridad")

        # 2. Aplicar el Motor de Reglas (Simula el OUTER APPLY)
        columnas_regla = ["id_regla", "nombre_cuenta", "codigo", "subcodigo"]

        # LA CORRECCIÓN CLAVE: Agregamos result_type="expand" para decirle a Pandas
        # que el resultado de la función se divida correctamente en múltiples columnas.
        df_sap[columnas_regla] = df_sap.apply(
            lambda fila: self.evaluar_regla_para_fila(fila, df_reglas),
            axis=1,
            result_type="expand",
        )

        # 3. Formatear Columnas Finales
        df_sap["fecha_contabilizacion_dt"] = pd.to_datetime(
            df_sap["fecha_contabilizacion"]
        )
        df_sap["mes"] = df_sap["fecha_contabilizacion_dt"].dt.month

        meses_es = {
            1: "ENERO",
            2: "FEBRERO",
            3: "MARZO",
            4: "ABRIL",
            5: "MAYO",
            6: "JUNIO",
            7: "JULIO",
            8: "AGOSTO",
            9: "SEPTIEMBRE",
            10: "OCTUBRE",
            11: "NOVIEMBRE",
            12: "DICIEMBRE",
        }
        df_sap["nmes"] = df_sap["mes"].map(meses_es)
        df_sap["tiene_regla"] = np.where(df_sap["id_regla"].isna(), "NO", "SI")
        df_sap["usuario_id"] = df_sap["usuario_id"].fillna("-")

        # Limpieza de nulos generales para JSON plano
        df_sap = df_sap.replace({np.nan: None})

        # Seleccionar columnas tal cual tu SELECT final de SQL
        columnas_finales = [
            "mes",
            "nmes",
            "fecha_contabilizacion",
            "fecha_documento",
            "numero_documento",
            "transaccion_id",
            "folio",
            "tipo_documento",
            "linea",
            "cuenta_asociada",
            "nombre_cuenta_asociada",
            "proveedor",
            "descripcion",
            "comentario_linea",
            "cuenta_contrapartida",
            "nombre_contrapartida",
            "referencia_1",
            "referencia_2",
            "referencia_3",
            "cargo_abono_ml",
            "cargo_abono_me",
            "usuario_id",
            "autor",
            "centro_costo",
            "centro_area",
            "nombre_area",
            "tiene_regla",
            "nombre_cuenta",
            "codigo",
            "subcodigo",
        ]

        df_final = df_sap[columnas_finales]

        return df_final.to_dict(orient="records")

    def export_libro_mayor_to_excel(self, start_date, end_date, account):

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        data_clasificada = self.get_libro_mayor_account(start_date, end_date, account)
        df = pd.DataFrame(data_clasificada)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "mes",
                    "nmes",
                    "fecha_contabilizacion",
                    "cuenta_asociada",
                    "cargo_abono_ml",
                ]
            )

        headers_mapping = {
            "mes": "N° Mes",
            "nmes": "Mes",
            "fecha_contabilizacion": "F. Contabilización",
            "fecha_documento": "F. Documento",
            "numero_documento": "N° Documento",
            "transaccion_id": "ID Trans.",
            "folio": "Folio SAP",
            "tipo_documento": "Tipo Documento",
            "linea": "Línea",
            "cuenta_asociada": "Cuenta Original",
            "nombre_cuenta_asociada": "Nombre Cuenta SAP",
            "proveedor": "Socio de Negocio / Proveedor",
            "descripcion": "Descripción / Memo",
            "comentario_linea": "Comentario Línea",
            "cuenta_contrapartida": "Cuenta Contra",
            "nombre_contrapartida": "Nombre Contrapartida",
            "referencia_1": "Referencia 1",
            "referencia_2": "Referencia 2",
            "referencia_3": "Referencia 3",
            "cargo_abono_ml": "Monto ML",
            "cargo_abono_me": "Monto ME",
            "usuario_id": "ID Usuario",
            "autor": "Autor / Usuario",
            "centro_costo": "Centro Costo",
            "centro_area": "Centro Área",
            "nombre_area": "Nombre Área",
            "tiene_regla": "Clasificado (Regla)",
            "nombre_cuenta": "Cuenta Destino (Clasificación)",
            "codigo": "Código General",
            "subcodigo": "Subcódigo",
        }

        columns_to_keep = [col for col in headers_mapping.keys() if col in df.columns]
        df_filtered = df[columns_to_keep]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libro Mayor Clasificado"
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(
            start_color="244061", end_color="244061", fill_type="solid"
        )
        zebra_fill = PatternFill(
            start_color="F2F5F8", end_color="F2F5F8", fill_type="solid"
        )

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10, color="000000")

        light_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for col_idx, tech_key in enumerate(columns_to_keep, 1):
            cell = ws.cell(row=1, column=col_idx, value=headers_mapping[tech_key])
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for row_idx, row_data in enumerate(df_filtered.itertuples(index=False), 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = light_border

                if row_idx % 2 == 0:
                    cell.fill = zebra_fill

                current_key = columns_to_keep[col_idx - 1]

                if "fecha" in current_key or current_key in [
                    "mes",
                    "nmes",
                    "tiene_regla",
                ]:
                    cell.alignment = Alignment(horizontal="center")
                elif "cargo_abono" in current_key:
                    cell.number_format = (
                        '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    )
                    cell.alignment = Alignment(horizontal="right")
                elif (
                    isinstance(val, (int, float))
                    and "id" not in current_key
                    and "numero" not in current_key
                ):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_libro_mayor_to_excel_v2(self, start_date, end_date, account):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # 1. Obtener y filtrar data (Igual que antes)
        data_clasificada = self.get_libro_mayor_account(start_date, end_date, account)
        df = pd.DataFrame(data_clasificada)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "mes",
                    "nmes",
                    "codigo",
                    "subcodigo",
                    "nombre_cuenta",
                    "proveedor",
                    "cargo_abono_ml",
                ]
            )

        headers_mapping = {
            "mes": "N° Mes",
            "nmes": "Mes",
            "fecha_contabilizacion": "F. Contabilización",
            "fecha_documento": "F. Documento",
            "numero_documento": "N° Documento",
            "transaccion_id": "ID Trans.",
            "folio": "Folio SAP",
            "tipo_documento": "Tipo Documento",
            "linea": "Línea",
            "cuenta_asociada": "Cuenta Original",
            "nombre_cuenta_asociada": "Nombre Cuenta SAP",
            "proveedor": "Socio de Negocio / Proveedor",
            "descripcion": "Descripción / Memo",
            "comentario_linea": "Comentario Línea",
            "cuenta_contrapartida": "Cuenta Contra",
            "nombre_contrapartida": "Nombre Contrapartida",
            "referencia_1": "Referencia 1",
            "referencia_2": "Referencia 2",
            "referencia_3": "Referencia 3",
            "cargo_abono_ml": "Monto ML",
            "cargo_abono_me": "Monto ME",
            "usuario_id": "ID Usuario",
            "autor": "Autor / Usuario",
            "centro_costo": "Centro Costo",
            "centro_area": "Centro Área",
            "nombre_area": "Nombre Área",
            "tiene_regla": "Clasificado (Regla)",
            "nombre_cuenta": "Cuenta Destino (Clasificación)",
            "codigo": "Código General",
            "subcodigo": "Subcódigo",
        }

        columns_to_keep = [col for col in headers_mapping.keys() if col in df.columns]
        df_filtered = df[columns_to_keep]

        # Inicializar el libro de Excel
        wb = openpyxl.Workbook()

        # Estilos Globales
        header_fill = PatternFill(
            start_color="244061", end_color="244061", fill_type="solid"
        )
        hierarchy_fill = PatternFill(
            start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
        )  # Azul claro para totales
        zebra_fill = PatternFill(
            start_color="F2F5F8", end_color="F2F5F8", fill_type="solid"
        )

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True, color="000000")
        data_font = Font(name="Calibri", size=10, color="000000")

        light_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # =========================================================================
        # PESTAÑA 1: DETALLE (La que ya tenías armada)
        # =========================================================================
        ws_detalle = wb.active
        ws_detalle.title = "Detalle Clasificado"
        ws_detalle.views.sheetView[0].showGridLines = True

        for col_idx, tech_key in enumerate(columns_to_keep, 1):
            cell = ws_detalle.cell(
                row=1, column=col_idx, value=headers_mapping[tech_key]
            )
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for row_idx, row_data in enumerate(df_filtered.itertuples(index=False), 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_detalle.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = light_border
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill

                current_key = columns_to_keep[col_idx - 1]
                if "fecha" in current_key or current_key in [
                    "mes",
                    "nmes",
                    "tiene_regla",
                ]:
                    cell.alignment = Alignment(horizontal="center")
                elif "cargo_abono" in current_key:
                    cell.number_format = (
                        '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    )
                elif isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")

        # =========================================================================
        # PESTAÑA 2: RESUMEN MATRIZ (ESTILO TABLA DINÁMICA CON AGRUPACIÓN REAL)
        # =========================================================================
        ws_resumen = wb.create_sheet(title="Resumen Ejecutivo")
        ws_resumen.views.sheetView[0].showGridLines = True

        # Asegurar que los montos sean numéricos antes de pivotar
        df["cargo_abono_ml"] = pd.to_numeric(
            df["cargo_abono_ml"], errors="coerce"
        ).fillna(0)

        # Armamos la tabla dinámica
        df_pivot = pd.pivot_table(
            df,
            values="cargo_abono_ml",
            index=["codigo", "subcodigo", "nombre_cuenta", "proveedor"],
            columns=["mes", "nmes"],
            aggfunc="sum",
        ).fillna(0)

        # Configurar los botones de agrupación de Excel en la parte SUPERIOR e IZQUIERDA
        ws_resumen.sheet_properties.outlinePr.summaryBelow = False

        # Combinar filas 1 y 2 para el título de la columna principal
        ws_resumen.cell(row=1, column=1, value="Concepto").font = bold_font
        ws_resumen.cell(row=1, column=1).alignment = Alignment(
            vertical="center", horizontal="left"
        )
        ws_resumen.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        columnas_meses = df_pivot.columns  # Lista de tuplas [(4, "ABRIL"), (5, "MAYO")]

        # Dibujar los meses en las columnas (Ahora inician en la columna 2, es decir, la B)
        for col_idx, (num_mes, nombre_mes) in enumerate(columnas_meses, 2):
            cell = ws_resumen.cell(row=1, column=col_idx, value=nombre_mes)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

            sub_cell = ws_resumen.cell(row=2, column=col_idx, value="IMPORTES SOLES")
            sub_cell.fill = header_fill
            sub_cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
            sub_cell.alignment = Alignment(horizontal="center")

        # Rellenar los datos en una ÚNICA COLUMNA (Columna A) con niveles de agrupación
        current_row = 3
        ultimo_codigo = None
        ultimo_subcodigo = None
        ultima_cuenta = None

        for jerarquia, montos in df_pivot.iterrows():
            cod, subcod, cuenta, prov = jerarquia

            # --- NIVEL 0: CÓDIGO GENERAL (Nivel superior, no se colapsa) ---
            if cod != ultimo_codigo:
                cell_a = ws_resumen.cell(row=current_row, column=1, value=cod)
                cell_a.font = bold_font
                cell_a.fill = hierarchy_fill

                # Totales del Código
                totales_cod = df_pivot.xs(cod, level="codigo").sum()
                for c_idx, total_val in enumerate(totales_cod, 2):
                    t_cell = ws_resumen.cell(
                        row=current_row, column=c_idx, value=total_val
                    )
                    t_cell.font = bold_font
                    t_cell.number_format = (
                        '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    )
                    t_cell.fill = hierarchy_fill

                ws_resumen.row_dimensions[current_row].outlineLevel = 0
                ultimo_codigo = cod
                current_row += 1

            # --- NIVEL 1: SUBCODIGO (Hijo de Código) ---
            if subcod != ultimo_subcodigo:
                cell_a = ws_resumen.cell(
                    row=current_row, column=1, value=f"  ⊞ {subcod}"
                )
                cell_a.font = bold_font

                totales_sub = df_pivot.xs(
                    (cod, subcod), level=["codigo", "subcodigo"]
                ).sum()
                for c_idx, total_val in enumerate(totales_sub, 2):
                    t_cell = ws_resumen.cell(
                        row=current_row, column=c_idx, value=total_val
                    )
                    t_cell.font = bold_font
                    t_cell.number_format = (
                        '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    )

                # Nivel 1 de agrupación (se cierra bajo el nivel 0)
                ws_resumen.row_dimensions[current_row].outlineLevel = 1
                ultimo_subcodigo = subcod
                current_row += 1

            # --- NIVEL 2: NOMBRE DE CUENTA (Hijo de Subcódigo) ---
            if cuenta != ultima_cuenta:
                cell_a = ws_resumen.cell(
                    row=current_row, column=1, value=f"    ▪ {cuenta}"
                )
                cell_a.font = data_font

                totales_cta = df_pivot.xs(
                    (cod, subcod, cuenta),
                    level=["codigo", "subcodigo", "nombre_cuenta"],
                ).sum()
                for c_idx, total_val in enumerate(totales_cta, 2):
                    t_cell = ws_resumen.cell(
                        row=current_row, column=c_idx, value=total_val
                    )
                    t_cell.number_format = (
                        '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                    )

                ws_resumen.row_dimensions[current_row].outlineLevel = 2
                ultima_cuenta = cuenta
                current_row += 1

            # --- NIVEL 3: PROVEEDOR (Detalle final profundo) ---
            cell_a = ws_resumen.cell(row=current_row, column=1, value=f"      {prov}")
            cell_a.font = data_font

            for c_idx, monto_mes in enumerate(montos, 2):
                m_cell = ws_resumen.cell(row=current_row, column=c_idx, value=monto_mes)
                m_cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                m_cell.border = light_border

            # El detalle más profundo pertenece al nivel 3
            ws_resumen.row_dimensions[current_row].outlineLevel = 3
            current_row += 1

        # Ajustar ancho de columnas para ambas hojas
        for sheet in [ws_detalle, ws_resumen]:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # 6. Guardar en memoria y retornar buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
